import os
import sys
import subprocess

import cv2
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


_HERE = os.path.dirname(os.path.abspath(__file__))


def _save_three_xlsx(sam2_rows, roi_rows, ir_rows, out_dir, inverse_rows=None):
    """Save the tracked temperature series into separate xlsx files."""
    if not _HAS_OPENPYXL:
        print("[Excel] openpyxl not installed; skip xlsx export")
        return

    def _make_wb(headers, rows, fill_color):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "frame_data"

        hdr_fill = PatternFill("solid", fgColor=fill_color)
        hdr_font = Font(bold=True, color="FFFFFF")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        valid_rows = []
        for row_idx, row in enumerate(rows, 2):
            clean = [
                value if not isinstance(value, float) or not np.isnan(value) else None
                for value in row
            ]
            for col_idx, value in enumerate(clean, 1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=round(value, 3) if isinstance(value, float) else value,
                )
            valid_rows.append(clean)

        ws2 = wb.create_sheet("summary")
        temps = [r[-1] for r in valid_rows if r[-1] is not None]
        times = [r[2] for r in valid_rows if r[-1] is not None]
        if temps:
            stats = [
                ("total_frames", len(rows)),
                ("valid_temp_frames", len(temps)),
                ("duration_s", round(rows[-1][2] - rows[0][2], 1) if rows else 0),
                ("temp_mean_c", round(float(np.mean(temps)), 2)),
                ("temp_max_c", round(float(np.max(temps)), 2)),
                ("temp_min_c", round(float(np.min(temps)), 2)),
                ("temp_std_c", round(float(np.std(temps)), 2)),
                ("peak_time_s", round(times[int(np.argmax(temps))], 1)),
            ]
        else:
            stats = [("total_frames", len(rows)), ("valid_temp_frames", 0)]

        ws2.cell(row=1, column=1, value="item").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="value").font = Font(bold=True)
        for row_idx, (key, value) in enumerate(stats, 2):
            ws2.cell(row=row_idx, column=1, value=key)
            ws2.cell(row=row_idx, column=2, value=value)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 14
        return wb, len(valid_rows)

    exports = [
        (
            "SAM2",
            ["frame_abs", "frame_rel", "time_s", "mask_pixels", "mask_pct", "temp_mean_c", "temp_min_c", "temp_max_c"],
            sam2_rows,
            "1F4E79",
            "temp_sam2.xlsx",
        ),
        (
            "ROI",
            ["frame_abs", "frame_rel", "time_s", "roi_temp_mean_c"],
            roi_rows,
            "833C00",
            "temp_roi.xlsx",
        ),
        (
            "IR",
            ["frame_abs", "frame_rel", "time_s", "ir_temp_mean_c"],
            ir_rows,
            "375623",
            "temp_ir.xlsx",
        ),
    ]

    for label, headers, rows, color, filename in exports:
        wb, row_count = _make_wb(headers, rows, color)
        path = os.path.join(out_dir, filename)
        wb.save(path)
        print(f"[Excel] {label:<5} saved: {path}  ({row_count} rows)")

    if inverse_rows:
        wb, row_count = _make_wb(
            ["frame_abs", "frame_rel", "time_s", "inverse_temp_mean_c"],
            inverse_rows,
            "4B2E84",
        )
        path = os.path.join(out_dir, "temp_inverse.xlsx")
        wb.save(path)
        print(f"[Excel] Inv   saved: {path}  ({row_count} rows)")


def _plot_three_curves(sam2_rows, roi_rows, ir_rows, out_path, inverse_rows=None):
    """Plot the exported temperature curves into one comparison image."""
    def _extract(rows, value_col):
        xs, ys = [], []
        for row in rows:
            value = row[value_col]
            if value is not None and not (isinstance(value, float) and np.isnan(value)):
                xs.append(row[2])
                ys.append(float(value))
        return xs, ys

    fig, ax = plt.subplots(figsize=(14, 4))

    t_sam2, v_sam2 = _extract(sam2_rows, 5)
    t_roi, v_roi = _extract(roi_rows, 3)
    t_ir, v_ir = _extract(ir_rows, 3)

    if t_sam2:
        t_min = [r[6] for r in sam2_rows if r[6] is not None and not np.isnan(r[6])]
        t_max = [r[7] for r in sam2_rows if r[7] is not None and not np.isnan(r[7])]
        if t_min and t_max and len(t_min) == len(t_sam2):
            ax.fill_between(t_sam2, t_min, t_max, alpha=0.15, color="#FF7F0E")
        ax.plot(t_sam2, v_sam2, color="#FF7F0E", lw=1.5, label="SAM2 Mask")
    if t_roi:
        ax.plot(t_roi, v_roi, color="#1F77B4", lw=1.5, label="ROI Fixed")
    if t_ir:
        ax.plot(t_ir, v_ir, color="#2CA02C", lw=1.5, label="IR Auto")
    if inverse_rows:
        t_inv, v_inv = _extract(inverse_rows, 3)
        if t_inv:
            ax.plot(t_inv, v_inv, color="#9B59B6", lw=1.5, label="Inverse (Wok-Bottom)")

    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Temperature (C)")
    ax.set_title(
        "Food Temperature - Four Strategies Comparison"
        if inverse_rows else
        "Food Temperature - Three Strategies Comparison"
    )
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"[Curve] saved: {out_path}")


def stitch_rgb_ir(
    rgb_viz_path,
    temp_data,
    ir_fps,
    wok_cfg,
    out_path,
    rgb_start_frame,
    rgb_fps,
    pct=40,
    wok_cx_history=None,
    inv_viz_path=None,
    info_h=50,
    chart_h=120,
):
    """Build the combined RGB/IR output video."""
    tools_dir = os.path.join(_HERE, "..", "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)
    from ir_mask_viz import render_ir_frame

    cap = cv2.VideoCapture(rgb_viz_path)
    if not cap.isOpened():
        print(f"[Stitch] failed to open RGB viz video: {rgb_viz_path}")
        return

    cap_inv = None
    if inv_viz_path and os.path.exists(inv_viz_path):
        cap_inv = cv2.VideoCapture(inv_viz_path)
        if not cap_inv.isOpened():
            cap_inv = None

    rgb_total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    out_fps = cap.get(cv2.CAP_PROP_FPS) or rgb_fps
    viz_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    viz_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    pure_h = viz_h - info_h - chart_h
    pure_w = viz_w
    n_ir = temp_data.shape[0]

    ir_aspect = temp_data.shape[2] / temp_data.shape[1]
    ir_out_h = pure_h
    ir_out_w = int(round(ir_out_h * ir_aspect))
    if ir_out_w % 2 != 0:
        ir_out_w += 1

    n_rgb_cols = 2 if cap_inv is not None else 1
    total_w = pure_w * n_rgb_cols + ir_out_w

    new_info_h = int(info_h * 1.8)
    new_chart_h = int(chart_h * 1.8)
    if new_info_h % 2 != 0:
        new_info_h += 1
    if new_chart_h % 2 != 0:
        new_chart_h += 1

    total_h = pure_h + new_info_h + new_chart_h

    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(out_path, fourcc, out_fps, (total_w, total_h))

    hist_sorted = sorted(wok_cx_history, key=lambda item: item[0]) if wok_cx_history else []
    hist_frames = [item[0] for item in hist_sorted]

    def _get_dynamic_wok(abs_frame):
        if not hist_sorted:
            return wok_cfg
        import bisect
        pos = bisect.bisect_right(hist_frames, abs_frame) - 1
        if pos < 0:
            return wok_cfg
        _, cx, cy = hist_sorted[pos]
        dyn = dict(wok_cfg)
        dyn["cx"] = cx
        dyn["cy"] = cy
        return dyn

    mode_str = "three-panel (RGB food + RGB inverse + IR)" if cap_inv else "two-panel (RGB + IR)"
    print(f"[Stitch] mode: {mode_str}")
    print(f"[Stitch] source: RGB={pure_w}x{pure_h}  IR={ir_out_w}x{ir_out_h}")
    print(f"[Stitch] bottom panels: info={new_info_h}px  chart={new_chart_h}px  width={total_w}px")
    print(f"[Stitch] output: {total_w}x{total_h}  total={rgb_total} frames")
    if hist_sorted:
        print(f"[Stitch] dynamic wok centers: {len(hist_sorted)} history records")

    for idx in range(rgb_total):
        ret, viz_frame = cap.read()
        if not ret:
            break

        pure_rgb = viz_frame[0:pure_h, :]
        info_src = viz_frame[pure_h:pure_h + info_h, :]
        chart_src = viz_frame[pure_h + info_h:pure_h + info_h + chart_h, :]

        inv_rgb = None
        if cap_inv is not None:
            ret_inv, viz_inv = cap_inv.read()
            inv_rgb = viz_inv[0:pure_h, :] if ret_inv else np.zeros_like(pure_rgb)

        time_s = (rgb_start_frame + idx) / rgb_fps
        ir_idx = min(int(round(time_s * ir_fps)), n_ir - 1)
        abs_frame = rgb_start_frame + idx
        dyn_wok = _get_dynamic_wok(abs_frame)
        ir_img = render_ir_frame(
            temp_data[ir_idx], dyn_wok, pct=pct, out_w=ir_out_w, out_h=ir_out_h
        )

        top_panel = np.hstack([pure_rgb, inv_rgb, ir_img]) if inv_rgb is not None else np.hstack([pure_rgb, ir_img])
        info_big = cv2.resize(info_src, (total_w, new_info_h), interpolation=cv2.INTER_LINEAR)
        chart_big = cv2.resize(chart_src, (total_w, new_chart_h), interpolation=cv2.INTER_LINEAR)

        writer.write(np.vstack([top_panel, info_big, chart_big]))

        if idx % 200 == 0:
            print(f"  stitch frame {idx}/{rgb_total}  ir={ir_idx}", end="\r")

    cap.release()
    if cap_inv is not None:
        cap_inv.release()
    writer.release()
    print(f"\n[Stitch] mp4v written: {out_path}")

    h264_path = out_path.replace(".mp4", "_h264.mp4")
    try:
        ffmpeg_cmd = [
            "ffmpeg", "-y",
            "-i", out_path,
            "-c:v", "libx264",
            "-preset", "fast",
            "-crf", "20",
            "-pix_fmt", "yuv420p",
            h264_path,
        ]
        print(f"[ffmpeg] convert to H.264: {os.path.basename(h264_path)}")
        result = subprocess.run(ffmpeg_cmd, capture_output=True, timeout=600)
        if result.returncode == 0 and os.path.exists(h264_path):
            os.replace(h264_path, out_path)
            print(f"[ffmpeg] conversion done, replaced original: {out_path}")
        else:
            err = result.stderr.decode("utf-8", errors="ignore")[-500:]
            print(f"[ffmpeg] conversion failed (returncode={result.returncode}), keep mp4v output")
            print(f"[ffmpeg] stderr: {err}")
    except FileNotFoundError:
        print("[ffmpeg] not found; skip H.264 conversion")
    except Exception as exc:
        print(f"[ffmpeg] conversion exception: {exc}")
