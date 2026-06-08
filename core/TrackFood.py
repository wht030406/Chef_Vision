"""
TrackFood.py — SAM2 视频追踪食材 + 温度融合（分批处理版）

流程：
1. 读取 food_labels.json（由 LabelFirstFrame.py 生成）
2. 将视频分批抽帧（每批 CHUNK_SIZE 帧），避免内存溢出
3. 用 SAM2 VideoPredictor 逐批追踪食材，批间传递 mask 状态
4. 每帧输出食材 mask
5. 用 homography.npy 将 mask 映射到红外图像坐标
6. 统计每帧菜区域的温度均值，输出 CSV + 可视化视频

依赖：
  pip install sam2 opencv-python numpy matplotlib
  权重：D:/sam2_checkpoints/sam2.1_hiera_large.pt
"""

import os
import sys
import json
import shutil
import glob
import re
from datetime import datetime

import cv2
import numpy as np
import torch
import matplotlib
matplotlib.use("Agg")   # 无头模式，不弹窗
import matplotlib.pyplot as plt

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── 路径基准（本文件所在目录）────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))

# ── 配置 ─────────────────────────────────────────────────────────────────────
LABELS_JSON     = os.path.join(_HERE, "food_labels.json")
HOMOGRAPHY_PATH = os.path.join(_HERE, "..", "data", "homography.npy")
OUTPUT_VIDEO    = os.path.join(_HERE, "..", "output", "track_result.mp4")
OUTPUT_VIDEO_VIZ = os.path.join(_HERE, "..", "output", "track_result_viz.mp4")
OUTPUT_CSV      = os.path.join(_HERE, "..", "output", "food_temp_log.csv")
OUTPUT_EXCEL    = os.path.join(_HERE, "..", "output", "food_temp_log.xlsx")

# SAM2 配置
# 模型选择：tiny（快，适合实时）/ large（慢，精度高，适合离线）
# MODEL_CFG       = "configs/sam2.1/sam2.1_hiera_t.yaml"
# CHECKPOINT_PATH = os.path.join(_HERE, "..", "models", "sam2.1_hiera_tiny.pt")
# 切换到 large：
MODEL_CFG       = "configs/sam2.1/sam2.1_hiera_l.yaml"
CHECKPOINT_PATH = os.path.join(_HERE, "..", "models", "sam2.1_hiera_large.pt")

# 分批处理参数
OPTICAL_FLOW_INTERVAL = 0    # 0 = 纯 SAM2 模式（推荐）；>1 = SAM2+光流混合

# SAM2 批次大小（帧数）
# 100 帧 = 4 秒批次（@25fps），流水线模式下处理时间(~3.5s) < 录制时间(4s)
CHUNK_SIZE      = 100

# SAM2 推理分辨率缩放
# 缩小分辨率可大幅提升 SAM2 速度，mask 会放大回原始分辨率用于温度计算
# None = 不缩放（使用原始分辨率）；(640, 480) = 缩放到 640×480
SAM2_INFER_SIZE = None   # SAM2 内部会 resize 到 1024p，外部缩放无效

# 提示点数量上限（None = 不限制，使用全部标注点）
MAX_FG_POINTS   = None
MAX_BG_POINTS   = None

# 温度数据（自动推断）
TEMP_NPY_PATH   = None   # None = 自动扫描项目目录

# 可视化参数
MASK_COLOR      = (0, 255, 100)   # BGR 绿色，食材 mask 叠加色
MASK_ALPHA      = 0.45
SHOW_PREVIEW    = False           # 关闭实时预览，减少资源占用

# 底部信息栏（可视化版输出）
INFO_H          = 50              # 文字信息条高度（像素）
CHART_H         = 120             # 温度曲线图高度（像素）
CURVE_WIN_S     = 60              # 曲线滑动窗口（秒），只显示最近 N 秒

# 分段自动重标点（模拟实时流水线）
# 每隔 N 秒自动重新生成前景点并重置 SAM2，解决长时间追踪漂移问题
# 0 = 关闭（使用 food_labels.json 中的关键帧）
RELABEL_INTERVAL_S = 8

# 临时帧目录（放在 core/ 下）
TMP_BASE        = os.path.join(_HERE, "tmp_sam2_frames")


# ── 工具函数 ─────────────────────────────────────────────────────────────────

def load_labels(path):
    """
    加载标注文件，兼容新格式（多关键帧列表）和旧格式（flat 单帧）。

    返回:
        video_path  : str
        start_frame : int           第一个关键帧的帧号（追踪起始点）
        keyframes   : list[dict]    所有关键帧，按 frame 升序排列
          每条: {"frame": int, "fg_points": [...], "bg_points": [...], "label": str}
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    print(f"[标注] 视频: {data['video_path']}")

    if "keyframes" in data:
        # ── 新格式 ────────────────────────────────────────────────────────────
        kfs = sorted(data["keyframes"], key=lambda k: k["frame"])
        for kf in kfs:
            fg = kf.get("fg_points", [])
            bg = kf.get("bg_points", [])
            kf["fg_points"] = fg if MAX_FG_POINTS is None else fg[:MAX_FG_POINTS]
            kf["bg_points"] = bg if MAX_BG_POINTS is None else bg[:MAX_BG_POINTS]
        print(f"[标注] 共 {len(kfs)} 个关键帧：")
        for kf in kfs:
            print(f"  帧 {kf['frame']:6d} ({kf['time_s']:.1f}s)  "
                  f"标签={kf.get('label','')}  "
                  f"FG={len(kf['fg_points'])}  BG={len(kf['bg_points'])}")
        start_frame = kfs[0]["frame"]
        return data["video_path"], start_frame, kfs
    else:
        # ── 旧格式兼容 ────────────────────────────────────────────────────────
        fg_all = data["fg_points"]
        bg_all = data["bg_points"]
        fg = fg_all if MAX_FG_POINTS is None else fg_all[:MAX_FG_POINTS]
        bg = bg_all if MAX_BG_POINTS is None else bg_all[:MAX_BG_POINTS]
        start_frame = data.get("start_frame", 0)
        fps_json    = data.get("fps", 25.0)
        print(f"[标注] 旧格式：起始帧={start_frame}  FG={len(fg)}  BG={len(bg)}")
        kf = {
            "frame":     start_frame,
            "time_s":    round(start_frame / fps_json, 3),
            "label":     "初始标注",
            "fg_points": fg,
            "bg_points": bg,
        }
        return data["video_path"], start_frame, [kf]


def find_temp_npy(video_path):
    """
    自动匹配温度 npy 文件。
    策略：
    1. 先找与视频同名的（rgb_ → temp_）
    2. 再扫描目录找所有 temp_*.npy，选时间戳最近的
    """
    base      = os.path.splitext(os.path.basename(video_path))[0]
    video_dir = os.path.dirname(os.path.abspath(video_path))

    # 策略 1：同名替换
    candidate = base.replace("rgb_", "temp_") + ".npy"
    candidate_path = os.path.join(video_dir, candidate)
    if os.path.exists(candidate_path):
        print(f"[温度] 找到同名温度文件: {candidate}")
        return candidate_path

    # 策略 2：提取视频时间戳，找最近的 temp_*.npy
    m = re.search(r"(\d{8}_\d{6})", base)
    if m:
        vid_ts = m.group(1)
        all_npy = glob.glob(os.path.join(video_dir, "temp_*.npy"))
        best_path, best_diff = None, float("inf")
        for p in all_npy:
            mn = re.search(r"(\d{8}_\d{6})", os.path.basename(p))
            if mn:
                diff = abs(int(mn.group(1).replace("_", "")) - int(vid_ts.replace("_", "")))
                if diff < best_diff:
                    best_diff, best_path = diff, p
        if best_path:
            print(f"[温度] 自动匹配到最近温度文件: {os.path.basename(best_path)}"
                  f"  (时间差 {best_diff})")
            return best_path

    print(f"[温度] 未找到匹配的温度文件，跳过温度统计")
    return None


def load_temp_data(npy_path):
    """
    加载完整温度矩阵，不做帧号切片。
    帧对齐由调用方根据 IR/RGB 帧率比例动态计算。
    返回 (data, ir_total_frames)
    """
    if npy_path is None or not os.path.exists(npy_path):
        return None, 0
    data = np.load(npy_path)
    print(f"[温度] 加载 {os.path.basename(npy_path)}，shape: {data.shape}，dtype: {data.dtype}")
    if data.ndim == 2:
        data = data[np.newaxis, ...]
    t_min_val = float(np.min(data))
    t_max_val = float(np.max(data))
    print(f"[温度] 总帧数: {data.shape[0]}  温度范围: {t_min_val:.1f}°C ~ {t_max_val:.1f}°C")
    return data, data.shape[0]


def build_sam2_predictor(device):
    from sam2.build_sam import build_sam2_video_predictor
    print(f"\n[SAM2] 加载模型: {CHECKPOINT_PATH}")
    predictor = build_sam2_video_predictor(MODEL_CFG, CHECKPOINT_PATH, device=device)
    print("[SAM2] 模型加载完成")
    return predictor


def extract_chunk_to_dir(video_path, start_abs, end_abs, infer_size=None):
    """
    将视频 [start_abs, end_abs) 帧抽到临时目录。
    infer_size: (W, H) 缩放尺寸，None 则不缩放。
    返回 (tmp_dir, frame_names, actual_count)
    """
    os.makedirs(TMP_BASE, exist_ok=True)
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix="chunk_", dir=TMP_BASE)

    cap = cv2.VideoCapture(video_path)
    cap.set(cv2.CAP_PROP_POS_FRAMES, start_abs)

    frame_names = []
    local_idx   = 0
    for _ in range(end_abs - start_abs):
        ret, frame = cap.read()
        if not ret:
            break
        if infer_size is not None:
            frame = cv2.resize(frame, infer_size, interpolation=cv2.INTER_AREA)
        fname = f"{local_idx:06d}.jpg"
        cv2.imwrite(os.path.join(tmp_dir, fname),
                    frame, [cv2.IMWRITE_JPEG_QUALITY, 92])
        frame_names.append(fname)
        local_idx += 1

    cap.release()
    return tmp_dir, frame_names, local_idx


def scale_points(points, src_wh, dst_wh):
    """将标注点从 src_wh 坐标系缩放到 dst_wh 坐标系"""
    if not points or src_wh == dst_wh:
        return points
    sx = dst_wh[0] / src_wh[0]
    sy = dst_wh[1] / src_wh[1]
    return [[p[0] * sx, p[1] * sy] for p in points]


def upscale_mask(mask, dst_wh):
    """将 mask 放大到 dst_wh (W, H) 大小"""
    mh, mw = mask.shape
    if (mw, mh) == dst_wh:
        return mask
    m_u8 = mask.astype(np.uint8) * 255
    m_up = cv2.resize(m_u8, dst_wh, interpolation=cv2.INTER_NEAREST)
    return m_up > 127


def track_chunk(predictor, tmp_dir, frame_names,
                fg_points, bg_points,
                carry_mask=None,
                inject_keyframes=None):
    """
    对一批帧进行 SAM2 追踪，支持在批内指定帧号注入新前景点。

    参数：
      fg_points / bg_points : 第一批（或无 carry_mask 时）使用的标注点
      carry_mask            : 上批末帧 mask (H,W bool)；None 表示第一批
      inject_keyframes      : list[dict]，本批内需要注入的额外关键帧，格式：
                              [{"local_frame": int, "fg_points": [...], "bg_points": [...]}]
                              local_frame 是在本批帧序列中的局部帧号（0-based）

    返回：
      masks    : {local_idx: mask (H,W bool)}
      last_mask: 本批末帧 mask（传给下一批）
    """
    inject_keyframes = inject_keyframes or []

    with torch.inference_mode(), torch.autocast("cuda", dtype=torch.bfloat16):
        inference_state = predictor.init_state(video_path=tmp_dir)

        if carry_mask is not None and carry_mask.any():
            # ── 后续批：用上批末帧 mask 传递边界（add_new_mask 与 points 在同帧互斥）
            predictor.add_new_mask(
                inference_state=inference_state,
                frame_idx=0,
                obj_id=1,
                mask=carry_mask,
            )
        else:
            # ── 第一批：用用户标注点 ────────────────────────────────────────
            points = np.array(fg_points + bg_points, dtype=np.float32)
            labels = np.array(
                [1] * len(fg_points) + [0] * len(bg_points),
                dtype=np.int32
            )
            predictor.add_new_points_or_box(
                inference_state=inference_state,
                frame_idx=0,
                obj_id=1,
                points=points,
                labels=labels,
            )

        # ── 注入额外关键帧的前景点 ───────────────────────────────────────────
        # local_frame=0：补强点追加到第0帧（与 carry_mask 同帧，SAM2 会合并两者）
        # local_frame>0：食材入锅等关键帧，在对应帧注入
        for kf_inject in inject_keyframes:
            local_f  = kf_inject["local_frame"]
            kf_fg    = kf_inject["fg_points"]
            kf_bg    = kf_inject.get("bg_points", [])
            if not kf_fg:
                continue
            if 0 <= local_f < len(frame_names):
                pts    = np.array(kf_fg + kf_bg, dtype=np.float32)
                lbls   = np.array([1]*len(kf_fg) + [0]*len(kf_bg), dtype=np.int32)
                predictor.add_new_points_or_box(
                    inference_state=inference_state,
                    frame_idx=local_f,
                    obj_id=1,
                    points=pts,
                    labels=lbls,
                )
                print(f"  [注入] 局部帧 {local_f}：FG={len(kf_fg)} BG={len(kf_bg)}"
                      f"  标签={kf_inject.get('label','')}")

        masks = {}
        for out_frame_idx, out_obj_ids, out_mask_logits in predictor.propagate_in_video(
            inference_state
        ):
            mask = (out_mask_logits[0] > 0.0).cpu().numpy().squeeze().astype(bool)
            masks[out_frame_idx] = mask

        predictor.reset_state(inference_state)
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    last_mask = masks.get(len(frame_names) - 1, None)
    return masks, last_mask


def map_mask_to_ir(rgb_mask, homography, ir_shape):
    """用单应矩阵将 RGB mask 映射到红外图像坐标系"""
    H_ir, W_ir = ir_shape
    ys, xs = np.where(rgb_mask)
    if len(xs) == 0:
        return np.zeros(ir_shape, dtype=bool)

    pts_rgb = np.stack([xs, ys, np.ones(len(xs))], axis=1).T  # (3, N)
    pts_ir  = homography @ pts_rgb
    pts_ir  = pts_ir[:2] / pts_ir[2]

    xi = np.round(pts_ir[0]).astype(int)
    yi = np.round(pts_ir[1]).astype(int)
    valid = (xi >= 0) & (xi < W_ir) & (yi >= 0) & (yi < H_ir)
    xi, yi = xi[valid], yi[valid]

    ir_mask = np.zeros(ir_shape, dtype=bool)
    ir_mask[yi, xi] = True
    return ir_mask


def flow_propagate_mask(prev_gray, cur_gray, prev_mask):
    """
    用 Farneback 稠密光流将上一帧 mask 传播到当前帧。

    原理：
      1. 计算 prev→cur 的稠密光流场 (dx, dy)
      2. 对 prev_mask 中每个前景像素，按光流偏移映射到 cur 坐标
      3. 对结果做形态学闭运算填补空洞

    参数：
      prev_gray : (H,W) uint8 灰度图（上一帧）
      cur_gray  : (H,W) uint8 灰度图（当前帧）
      prev_mask : (H,W) bool  上一帧 mask

    返回：
      cur_mask  : (H,W) bool  传播后的 mask
    """
    if not prev_mask.any():
        return prev_mask.copy()

    # Farneback 光流（CPU，速度快）
    flow = cv2.calcOpticalFlowFarneback(
        prev_gray, cur_gray,
        None,
        pyr_scale=0.5, levels=3, winsize=15,
        iterations=3, poly_n=5, poly_sigma=1.2,
        flags=0
    )  # shape: (H, W, 2)

    H, W = prev_mask.shape
    ys, xs = np.where(prev_mask)

    # 按光流偏移计算新坐标
    dx = flow[ys, xs, 0]
    dy = flow[ys, xs, 1]
    new_xs = np.round(xs + dx).astype(int)
    new_ys = np.round(ys + dy).astype(int)

    # 过滤越界坐标
    valid = (new_xs >= 0) & (new_xs < W) & (new_ys >= 0) & (new_ys < H)
    new_xs = new_xs[valid]
    new_ys = new_ys[valid]

    cur_mask = np.zeros((H, W), dtype=np.uint8)
    cur_mask[new_ys, new_xs] = 255

    # 形态学闭运算：填补光流稀疏导致的空洞
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    cur_mask = cv2.morphologyEx(cur_mask, cv2.MORPH_CLOSE, kernel)

    return cur_mask.astype(bool)


def render_overlay(frame_bgr, mask, color_bgr, alpha):
    """在帧上叠加半透明 mask + 轮廓"""
    vis = frame_bgr.copy()
    c   = np.array(color_bgr, dtype=np.uint8)
    vis[mask] = (vis[mask].astype(float) * (1 - alpha) + c * alpha).astype(np.uint8)
    mask_u8 = mask.astype(np.uint8) * 255
    contours, _ = cv2.findContours(mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(vis, contours, -1, (255, 255, 255), 1)
    return vis


def _kmeans_food_temp(wok_temps, min_cluster_gap=30.0):
    """
    用 K-means 把锅内温度分成两类（高温=锅壁/锅底，低温=食物），
    返回低温类的均值作为食物温度。

    参数：
      wok_temps       : 1D float array，锅内所有像素温度
      min_cluster_gap : 两个聚类中心差距的最小值（°C）。
                        若两类中心差距 < 此值，说明锅内温度分布均匀
                        （锅倾斜/空锅/翻炒过渡期），返回 nan 表示帧不可靠。

    返回：
      float，食物温度均值；或 nan（帧不可靠）
    """
    vals = wok_temps.astype(np.float32).flatten()
    if len(vals) < 10:
        return float("nan")

    # 初始化：用最小值和最大值作为两个初始中心
    c_low  = float(np.percentile(vals, 10))
    c_high = float(np.percentile(vals, 90))

    # 迭代 K-means（最多 20 次，收敛即停）
    for _ in range(20):
        # 分配：每个像素归入更近的中心
        dist_low  = np.abs(vals - c_low)
        dist_high = np.abs(vals - c_high)
        label_low = dist_low <= dist_high   # True = 低温类（食物）

        new_low  = float(np.mean(vals[label_low]))  if label_low.any()  else c_low
        new_high = float(np.mean(vals[~label_low])) if (~label_low).any() else c_high

        if abs(new_low - c_low) < 0.1 and abs(new_high - c_high) < 0.1:
            break
        c_low, c_high = new_low, new_high

    # 可靠性检查：两类中心差距太小 → 帧不可靠（翻炒/倾斜/空锅）
    if (c_high - c_low) < min_cluster_gap:
        return float("nan")

    # 返回低温类（食物）的均值
    dist_low  = np.abs(vals - c_low)
    dist_high = np.abs(vals - c_high)
    food_mask = dist_low <= dist_high
    return float(np.mean(vals[food_mask])) if food_mask.any() else float("nan")


def draw_temp_chart(temp_history, cur_time_s, w, h, curve_win_s=60,
                    roi_history=None, ir_mask_history=None):
    """
    用纯 numpy/cv2 绘制温度折线图，支持三条曲线：
      SAM2 mask（橙色）、ROI 固定圆圈（蓝色）、IR mask 自动分割（绿色）

    参数：
      temp_history    : list of (time_s, temp_mean)，SAM2 mask 温度历史
      cur_time_s      : 当前帧时间（秒）
      w, h            : 图像宽高（像素）
      curve_win_s     : 滑动窗口长度（秒），只显示最近 N 秒
      roi_history     : list of (time_s, temp_mean)，ROI 区域温度历史（可选）
      ir_mask_history : list of (time_s, temp_mean)，IR mask 温度历史（可选）
    """
    bar = np.zeros((h, w, 3), dtype=np.uint8)
    if len(temp_history) < 2:
        cv2.putText(bar, "Mask Avg Temp (waiting for data...)",
                    (10, h // 2), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (160, 160, 160), 1)
        return bar

    # 取滑动窗口内的有效数据
    t0  = max(0.0, cur_time_s - curve_win_s)
    pts = [(t, v) for t, v in temp_history if t >= t0 and not np.isnan(v)]
    if len(pts) < 2:
        pts = temp_history[-2:]

    times = [p[0] for p in pts]
    vals  = [p[1] for p in pts]

    # ROI 数据（同窗口）
    roi_pts = []
    if roi_history:
        roi_pts = [(t, v) for t, v in roi_history if t >= t0 and not np.isnan(v)]

    # 坐标范围（Y 轴留 5°C 余量，兼顾两条曲线）
    all_vals = vals + [v for _, v in roi_pts]
    t_min = t0
    t_max = max(cur_time_s, t0 + 1.0)
    v_min = max(0.0, min(all_vals) - 5.0)
    v_max = max(all_vals) + 5.0
    if v_max <= v_min:
        v_max = v_min + 10.0

    # 绘图边距
    pad_l, pad_r, pad_t, pad_b = 48, 12, 10, 22

    def tx(t):
        return pad_l + int((t - t_min) / (t_max - t_min) * (w - pad_l - pad_r))

    def ty(v):
        return pad_t + int((1.0 - (v - v_min) / (v_max - v_min)) * (h - pad_t - pad_b))

    # 网格线 + Y 轴刻度（3 条）
    for v in np.linspace(v_min, v_max, 3):
        yy = ty(v)
        cv2.line(bar, (pad_l, yy), (w - pad_r, yy), (45, 45, 45), 1)
        cv2.putText(bar, f"{v:.0f}", (2, yy + 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.35, (180, 180, 180), 1)

    # 坐标轴
    cv2.line(bar, (pad_l, pad_t), (pad_l, h - pad_b), (160, 160, 160), 1)
    cv2.line(bar, (pad_l, h - pad_b), (w - pad_r, h - pad_b), (160, 160, 160), 1)

    # X 轴时间刻度（起止）
    cv2.putText(bar, f"{t_min:.0f}s", (pad_l, h - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.32, (140, 140, 140), 1)
    cv2.putText(bar, f"{cur_time_s:.1f}s", (w - pad_r - 30, h - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.32, (140, 140, 140), 1)

    # SAM2 mask 折线（橙色）
    screen_pts = [(tx(t), ty(v)) for t, v in zip(times, vals)]
    for i in range(1, len(screen_pts)):
        p1, p2 = screen_pts[i - 1], screen_pts[i]
        if all(0 <= c < dim for c, dim in [(p1[0], w), (p1[1], h), (p2[0], w), (p2[1], h)]):
            cv2.line(bar, p1, p2, (50, 165, 255), 2)

    # ROI 折线（蓝色）
    if len(roi_pts) >= 2:
        roi_screen = [(tx(t), ty(v)) for t, v in roi_pts]
        for i in range(1, len(roi_screen)):
            p1, p2 = roi_screen[i - 1], roi_screen[i]
            if all(0 <= c < dim for c, dim in [(p1[0], w), (p1[1], h), (p2[0], w), (p2[1], h)]):
                cv2.line(bar, p1, p2, (255, 160, 30), 2)
        # ROI 当前值
        rx, ry = roi_screen[-1]
        if 0 <= rx < w and 0 <= ry < h:
            cv2.circle(bar, (rx, ry), 4, (255, 100, 0), -1)
            cv2.putText(bar, f"ROI:{roi_pts[-1][1]:.1f}C", (rx + 6, ry + 4),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, (255, 180, 60), 1)

    # IR mask 折线（绿色）
    ir_pts = []
    if ir_mask_history:
        ir_pts = [(t, v) for t, v in ir_mask_history if t >= t0 and not np.isnan(v)]
    if len(ir_pts) >= 2:
        ir_screen = [(tx(t), ty(v)) for t, v in ir_pts]
        for i in range(1, len(ir_screen)):
            p1, p2 = ir_screen[i - 1], ir_screen[i]
            if all(0 <= c < dim for c, dim in [(p1[0], w), (p1[1], h), (p2[0], w), (p2[1], h)]):
                cv2.line(bar, p1, p2, (0, 220, 80), 2)
        irx, iry = ir_screen[-1]
        if 0 <= irx < w and 0 <= iry < h:
            cv2.circle(bar, (irx, iry), 4, (0, 255, 60), -1)
            cv2.putText(bar, f"IR:{ir_pts[-1][1]:.1f}C", (irx + 6, iry + 4),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 255, 100), 1)

    # SAM2 当前帧红点
    cx, cy = tx(cur_time_s), ty(vals[-1])
    if 0 <= cx < w and 0 <= cy < h:
        cv2.circle(bar, (cx, cy), 4, (0, 60, 255), -1)
        cv2.putText(bar, f"Mask:{vals[-1]:.1f}C", (cx + 6, cy + 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 200, 255), 1)

    # 图例（全英文，cv2 不支持中文/特殊符号）
    cv2.putText(bar, "[SAM2]", (pad_l + 4, pad_t + 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.38, (50, 165, 255), 1)
    if roi_pts:
        cv2.putText(bar, "[ROI]", (pad_l + 70, pad_t + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, (255, 160, 30), 1)
    if ir_pts:
        cv2.putText(bar, "[IR-Auto]", (pad_l + 120, pad_t + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 220, 80), 1)

    return bar


# ── 主程序 ───────────────────────────────────────────────────────────────────

def main():
    # ── 创建时间戳输出子目录 ──────────────────────────────────────────────────
    run_ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir  = os.path.join(_HERE, "..", "output", run_ts)
    os.makedirs(out_dir, exist_ok=True)
    out_video_viz = os.path.join(out_dir, "track_result_viz.mp4")
    out_csv       = os.path.join(out_dir, "food_temp_log.csv")
    out_xlsx      = os.path.join(out_dir, "food_temp_log.xlsx")
    out_curve     = os.path.join(out_dir, "food_temp_curve.png")
    print(f"[输出] 本次结果目录: {out_dir}")

    # ── 检查依赖文件 ──────────────────────────────────────────────────────────
    if not os.path.exists(LABELS_JSON):
        print(f"[错误] 找不到标注文件: {LABELS_JSON}")
        print("请先运行 LabelFirstFrame.py 完成标注")
        sys.exit(1)

    video_path, start_frame, keyframes = load_labels(LABELS_JSON)
    # 第一个关键帧用于初始标注
    first_kf  = keyframes[0]
    fg_points = first_kf["fg_points"]
    bg_points = first_kf["bg_points"]
    # 其余关键帧（index>=1）将在追踪过程中按帧号注入
    extra_kfs = keyframes[1:]
    if extra_kfs:
        print(f"[多关键帧] 将在以下帧注入额外前景标注：")
        for kf in extra_kfs:
            print(f"  帧 {kf['frame']}  标签={kf.get('label','')}  FG={len(kf['fg_points'])}")

    if not os.path.exists(video_path):
        print(f"[错误] 找不到视频: {video_path}")
        sys.exit(1)

    # 获取视频基本信息
    cap_info = cv2.VideoCapture(video_path)
    total_frames = int(cap_info.get(cv2.CAP_PROP_FRAME_COUNT))
    fps          = cap_info.get(cv2.CAP_PROP_FPS)
    VW           = int(cap_info.get(cv2.CAP_PROP_FRAME_WIDTH))
    VH           = int(cap_info.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap_info.release()
    track_frames = total_frames - start_frame
    print(f"\n[视频] {video_path}")
    print(f"[视频] 分辨率: {VW}x{VH}  总帧数: {total_frames}  FPS: {fps:.1f}")
    print(f"[视频] 追踪范围: 第 {start_frame} ~ {total_frames} 帧，共 {track_frames} 帧")
    print(f"[分批] 每批 {CHUNK_SIZE} 帧，共需 {(track_frames + CHUNK_SIZE - 1)//CHUNK_SIZE} 批")

    # ── 预构建关键帧注入表：{abs_frame: kf_dict} ──────────────────────────────
    # 只包含 extra_kfs（index>=1），第一个关键帧已经作为初始标注点使用
    inject_map = {kf["frame"]: kf for kf in extra_kfs}

    # ── 加载单应矩阵（可选）──────────────────────────────────────────────────
    homography = None
    if os.path.exists(HOMOGRAPHY_PATH):
        homography = np.load(HOMOGRAPHY_PATH)
        print(f"[单应矩阵] 已加载: {HOMOGRAPHY_PATH}  shape: {homography.shape}")
    else:
        print(f"[单应矩阵] 未找到 {HOMOGRAPHY_PATH}，跳过温度融合")

    # ── 自动匹配温度文件 ──────────────────────────────────────────────────────
    global TEMP_NPY_PATH
    if TEMP_NPY_PATH is None:
        TEMP_NPY_PATH = find_temp_npy(video_path)
    temp_data, ir_total_frames = load_temp_data(TEMP_NPY_PATH)

    # 计算 IR/RGB 帧率比例（用于时间对齐）
    # 两路流同时开始录制，用各自总帧数推算帧率比
    ir_fps_ratio = 1.0   # 默认 1:1
    if temp_data is not None and total_frames > 0:
        ir_fps_ratio = ir_total_frames / total_frames
        ir_fps_est   = fps * ir_fps_ratio
        print(f"[帧率对齐] RGB {fps:.1f}fps × {total_frames}帧 | "
              f"IR ~{ir_fps_est:.1f}fps × {ir_total_frames}帧 | "
              f"比例 {ir_fps_ratio:.4f}")

    # ── 设备 ──────────────────────────────────────────────────────────────────
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\n[设备] 使用: {device}")
    if device.type == "cuda":
        print(f"[设备] GPU: {torch.cuda.get_device_name(0)}")
        # 限制 PyTorch 显存碎片，降低 OOM 风险
        os.environ.setdefault("PYTORCH_CUDA_ALLOC_CONF", "expandable_segments:True")

    # ── 加载 SAM2 ─────────────────────────────────────────────────────────────
    predictor = build_sam2_predictor(device)

    # ── 输出视频准备 ──────────────────────────────────────────────────────────
    fourcc   = cv2.VideoWriter_fourcc(*"mp4v")
    OUT_H    = VH + INFO_H + CHART_H          # 原始帧高 + 文字条 + 曲线图
    writer   = cv2.VideoWriter(out_video_viz, fourcc, fps, (VW, OUT_H))
    # 三策略逐帧数据（各自独立存储，输出为单独表格）
    sam2_rows   = []   # [frame_abs, frame_rel, time_s, mask_pixels, mask_ratio, mean, min, max]
    roi_rows    = []   # [frame_abs, frame_rel, time_s, roi_temp_mean]
    ir_rows     = []   # [frame_abs, frame_rel, time_s, ir_mask_temp]
    temp_history     = []   # list of (time_s, temp_mean)，SAM2 mask 温度历史
    roi_history      = []   # list of (time_s, roi_temp)，ROI 区域温度历史
    ir_mask_history  = []   # list of (time_s, ir_mask_temp)，IR 自动分割温度历史

    # ── 加载 ROI 配置 ─────────────────────────────────────────────────────────
    roi_cfg = None
    roi_cfg_path = os.path.join(_HERE, "..", "field", "roi_config.json")
    # 也兼容放在 data/ 目录下的情况
    if not os.path.exists(roi_cfg_path):
        roi_cfg_path = os.path.join(_HERE, "..", "data", "roi_config.json")
    if os.path.exists(roi_cfg_path):
        with open(roi_cfg_path, "r") as f:
            roi_cfg = json.load(f)
        print(f"[ROI] 已加载配置: {roi_cfg_path}")
        print(f"  RGB坐标: 圆心=({roi_cfg['rgb_cx']},{roi_cfg['rgb_cy']}) 半径={roi_cfg['rgb_radius']}px")
    else:
        print(f"[ROI] 未找到 roi_config.json，跳过 ROI 温度统计")
        print(f"  提示：在 FieldCapture.py 中按 R 键设置 ROI 后会自动生成")

    # ── 加载 IR 锅区域配置（用于 IR 自动分割温度曲线）────────────────────────
    wok_cfg     = None
    wok_mask_ir = None
    IR_FOOD_PCT = 40   # 锅内低于此百分位的像素 = 菜
    wok_cfg_path = os.path.join(_HERE, "..", "data", "wok_region.json")
    if os.path.exists(wok_cfg_path) and temp_data is not None:
        with open(wok_cfg_path) as f:
            wok_cfg = json.load(f)
        ir_h_wok = temp_data.shape[1]
        ir_w_wok = temp_data.shape[2]
        wok_mask_ir = np.zeros((ir_h_wok, ir_w_wok), dtype=np.uint8)
        cv2.ellipse(wok_mask_ir,
                    (wok_cfg["cx"], wok_cfg["cy"]),
                    (wok_cfg["rx"], wok_cfg["ry"]),
                    0, 0, 360, 255, -1)
        wok_mask_ir = wok_mask_ir > 0
        print(f"[IR Mask] 已加载锅区域: {wok_cfg_path}")
        print(f"  cx={wok_cfg['cx']} cy={wok_cfg['cy']} "
              f"rx={wok_cfg['rx']} ry={wok_cfg['ry']}  "
              f"覆盖 {wok_mask_ir.sum()} 像素")
    else:
        print(f"[IR Mask] 未找到 wok_region.json 或无温度数据，跳过 IR 自动分割温度")

    # ── 分批追踪主循环（SAM2 + 光流混合）────────────────────────────────────
    # 真正的混合模式：
    #   - 纯SAM2模式（OPTICAL_FLOW_INTERVAL<=1）：每批处理 CHUNK_SIZE 帧
    #   - 混合模式（OPTICAL_FLOW_INTERVAL>1）：SAM2 只处理锚点帧（单帧批次），
    #     其余帧完全用光流传播，不再调用 SAM2
    carry_mask     = None   # 上批末帧 SAM2 mask，用于跨批传递
    global_local   = 0      # 全局相对帧计数
    flow_prev_gray = None   # 光流：上一帧灰度图
    flow_prev_mask = None   # 光流：上一帧 mask

    use_flow = (OPTICAL_FLOW_INTERVAL > 1)
    if use_flow:
        print(f"[模式] SAM2+光流混合  SAM2锚点间隔={OPTICAL_FLOW_INTERVAL}帧")
        # 混合模式下：SAM2 每隔 N 帧处理 1 帧，其余帧光流传播
        # 将追踪范围按锚点帧切分
        anchor_frames = list(range(start_frame,
                                   total_frames,
                                   OPTICAL_FLOW_INTERVAL))
        print(f"[混合] 共 {len(anchor_frames)} 个SAM2锚点帧，"
              f"其余 {track_frames - len(anchor_frames)} 帧用光流")
    else:
        print(f"[模式] 纯SAM2  CHUNK_SIZE={CHUNK_SIZE}")

    # ── 推理分辨率：计算缩放比例 ─────────────────────────────────────────────
    orig_wh   = (VW, VH)   # 原始分辨率
    infer_wh  = SAM2_INFER_SIZE if SAM2_INFER_SIZE is not None else orig_wh
    do_resize = (infer_wh != orig_wh)

    if do_resize:
        print(f"[缩放] SAM2推理: {infer_wh[0]}×{infer_wh[1]}  "
              f"输出/温度: {orig_wh[0]}×{orig_wh[1]}")
        # 将标注点坐标缩放到推理分辨率
        fg_infer = scale_points(fg_points, orig_wh, infer_wh)
        bg_infer = scale_points(bg_points, orig_wh, infer_wh)
    else:
        fg_infer = fg_points
        bg_infer = bg_points

    # ── 自动重标点：加载 auto_label 的标点函数 ────────────────────────────────
    _auto_label_func = None
    _wok_cfg_al      = None
    _wok_mask_al     = None
    _H_inv_al        = None
    _rng_al          = None
    if RELABEL_INTERVAL_S > 0:
        try:
            import sys as _sys
            if _HERE not in _sys.path:
                _sys.path.insert(0, _HERE)
            from auto_label import (
                load_wok_cfg as _al_load_wok,
                build_wok_mask as _al_build_mask,
                generate_ir_mask_and_points as _al_ir_mask,
            )
            _wok_cfg_al = _al_load_wok()
            if temp_data is not None and homography is not None:
                _ir_h_al = temp_data.shape[1]
                _ir_w_al = temp_data.shape[2]
                _wok_mask_al = _al_build_mask(_wok_cfg_al, _ir_h_al, _ir_w_al)
                _H_inv_al    = np.linalg.inv(homography)
                _rng_al      = np.random.default_rng(42)
                _auto_label_func = _al_ir_mask
                print(f"[自动重标点] 已加载，每 {RELABEL_INTERVAL_S}s 重新标点并重置SAM2")
            else:
                print(f"[自动重标点] 无温度数据或单应矩阵，禁用")
        except ImportError as e:
            print(f"[自动重标点] 导入失败({e})，禁用")

    last_relabel_s = start_frame / fps   # 上次重标点的时间（秒）
    prev_mask_ratio = 0.0                # 上批末帧 mask 面积占比（%），用于异常检测
    last_reinforce_wok_pct = 0.0         # 上次补强时 mask 占 wok 区域的%（骤降检测用）

    # ── 预计算 wok RGB mask（用于每帧后处理约束，避免循环内重复计算）────────
    _wok_rgb_constraint = None   # (VH, VW) bool
    if wok_mask_ir is not None and homography is not None:
        try:
            _H_inv_wok = np.linalg.inv(homography)
            _wok_u8_pre = wok_mask_ir.astype(np.uint8) * 255
            _wok_proj   = cv2.warpPerspective(_wok_u8_pre, _H_inv_wok, (VW, VH))
            _wok_rgb_constraint = _wok_proj > 64
            print(f"[wok约束] 预计算 RGB锅区域 mask  覆盖像素: {_wok_rgb_constraint.sum()}")
        except Exception as _we:
            print(f"[wok约束] 预计算失败({_we})，跳过约束")

    # ── 保存初始关键帧预览图 ──────────────────────────────────────────────────
    if _auto_label_func is not None:
        try:
            _cap_init = cv2.VideoCapture(video_path)
            _cap_init.set(cv2.CAP_PROP_POS_FRAMES, start_frame)
            _ret_init, _rgb_init = _cap_init.read()
            _cap_init.release()
            if _ret_init:
                _vis_init = _rgb_init.copy()
                for _p in fg_points:
                    cv2.circle(_vis_init, (int(_p[0]), int(_p[1])), 8, (0, 255, 80), -1)
                    cv2.circle(_vis_init, (int(_p[0]), int(_p[1])), 9, (255, 255, 255), 1)
                for _p in bg_points:
                    cv2.circle(_vis_init, (int(_p[0]), int(_p[1])), 8, (0, 0, 255), -1)
                    cv2.circle(_vis_init, (int(_p[0]), int(_p[1])), 9, (255, 255, 255), 1)
                cv2.putText(_vis_init,
                            f"Initial-Label t={start_frame/fps:.0f}s  FG(green):{len(fg_points)} BG(red):{len(bg_points)}",
                            (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 200, 255), 2)
                cv2.putText(_vis_init, f"[from food_labels.json]  frame={start_frame}",
                            (20, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 160, 200), 2)
                _init_name = f"relabel_t{start_frame/fps:.0f}s_f{start_frame}_initial.jpg"
                cv2.imwrite(os.path.join(out_dir, _init_name), _vis_init)
                print(f"[初始标注] 预览图已保存: {_init_name}")
        except Exception as _ei:
            print(f"[初始标注] 预览图保存失败: {_ei}")

    if not use_flow:
        # ── 纯 SAM2 模式 ──────────────────────────────────────────────────────
        n_chunks = (track_frames + CHUNK_SIZE - 1) // CHUNK_SIZE
        for chunk_i in range(n_chunks):
            chunk_start_abs = start_frame + chunk_i * CHUNK_SIZE
            chunk_end_abs   = min(chunk_start_abs + CHUNK_SIZE, total_frames)
            chunk_len       = chunk_end_abs - chunk_start_abs
            chunk_start_s   = chunk_start_abs / fps

            print(f"\n{'='*55}")
            print(f"[批次 {chunk_i+1}/{n_chunks}] 帧 {chunk_start_abs} ~ {chunk_end_abs-1}"
                  f"  ({chunk_len} 帧)")

            # ── 检查是否需要自动补强（每 N 秒从 carry_mask 内部采点注入第一帧）────
            # 新策略：carry_mask 始终保留（不置 None），只用 SAM2 自己的末帧 mask 定位，
            # IR 只做稳定性检查（决定要不要补强），不再参与任何标注点生成。
            _reinforce_inject = None   # 本批第一帧要注入的补强点
            if (carry_mask is not None
                    and chunk_i > 0
                    and (chunk_start_s - last_relabel_s) >= RELABEL_INTERVAL_S):
                # ── IR 稳定性检查（只判断要不要补强，不采点）─────────────────
                _do_reinforce = True
                if _auto_label_func is not None and temp_data is not None:
                    try:
                        _ir_idx_al = int(chunk_start_abs * ir_fps_ratio)
                        _ir_idx_al = min(_ir_idx_al, temp_data.shape[0] - 1)
                        _ir_frame_al = temp_data[_ir_idx_al]
                        from auto_label import load_wok_cfg as _lwc, build_wok_mask as _bwm
                        _wok_temps_chk = _ir_frame_al[_wok_mask_al]
                        _var_chk = float(np.var(_wok_temps_chk)) if len(_wok_temps_chk) > 0 else 0
                        if _var_chk < 200.0:
                            _do_reinforce = False
                            print(f"[补强] t={chunk_start_s:.1f}s  锅倾斜/翻炒(var={_var_chk:.1f})，跳过")
                    except Exception:
                        pass   # 检查失败则默认补强

                if _do_reinforce:
                    # ── 检查 carry_mask 是否还在锅内（位置检查，比面积检查更可靠）────
                    # 指标：mask 与锅内区域(wok_rgb_constraint)的交集 / mask 自身面积
                    # < 60% 说明 mask 大部分跑到锅外/搅拌爪/金属部件上 → 强制重置
                    _mask_px = int(carry_mask.sum())
                    _need_reset = False
                    _overlap_pct = 100.0   # 默认100%（无wok约束时不重置）
                    if _wok_rgb_constraint is not None and _mask_px > 0:
                        _overlap_px = int((carry_mask & _wok_rgb_constraint).sum())
                        _overlap_pct = _overlap_px / _mask_px * 100
                        if _overlap_pct < 60.0:
                            _need_reset = True
                            print(f"[补强] t={chunk_start_s:.1f}s  "
                                  f"mask偏离锅内(overlap={_overlap_pct:.0f}%<60%)，"
                                  f"重置SAM2→初始标注点")
                    # 也保留面积检查：mask 超过 wok 区域 35% 同样重置（整锅漂移/锅底漂移）
                    # 食材正常情况下占锅内面积 < 35%，超过说明 SAM2 追踪到了锅底/空白区域
                    _wok_px_chk = int(_wok_rgb_constraint.sum()) if _wok_rgb_constraint is not None else (VW * VH)
                    _mask_vs_wok = _mask_px / max(_wok_px_chk, 1) * 100
                    if not _need_reset:
                        if _mask_vs_wok > 35.0:
                            _need_reset = True
                            print(f"[补强] t={chunk_start_s:.1f}s  "
                                  f"mask过大({_mask_vs_wok:.0f}%>wok35%)，"
                                  f"重置SAM2→IR定位新前景点")
                    # ── 第三级：面积骤降检测（漂移到旋转轴/小碎片）────────────
                    # last_reinforce_wok_pct = 上次补强时 mask 占 wok% （不是批次均值）
                    # 骤降 > 70% 或绝对值 < 2% → 重置
                    if not _need_reset and _wok_px_chk > 0:
                        _drop_pct = (last_reinforce_wok_pct - _mask_vs_wok) / max(last_reinforce_wok_pct, 0.1) * 100
                        if _mask_vs_wok < 2.0:
                            _need_reset = True
                            print(f"[补强] t={chunk_start_s:.1f}s  "
                                  f"mask过小({_mask_vs_wok:.1f}%<2%，可能是旋转轴碎片)，"
                                  f"重置SAM2→初始标注点")
                        elif last_reinforce_wok_pct > 5.0 and _drop_pct > 70.0:
                            _need_reset = True
                            print(f"[补强] t={chunk_start_s:.1f}s  "
                                  f"mask面积骤降({last_reinforce_wok_pct:.0f}%→{_mask_vs_wok:.0f}%，"
                                  f"跌幅{_drop_pct:.0f}%>70%)，重置SAM2→初始标注点")
                    if _need_reset:
                        # ── 先保存重置预览图（用跑偏的 carry_mask 可视化，方便排查）──
                        try:
                            _cap_rst = cv2.VideoCapture(video_path)
                            _cap_rst.set(cv2.CAP_PROP_POS_FRAMES, chunk_start_abs)
                            _ret_rst, _rgb_rst = _cap_rst.read()
                            _cap_rst.release()
                            if _ret_rst:
                                _vis_rst = _rgb_rst.copy()
                                # 用红色半透明显示跑偏的 carry_mask
                                if carry_mask is not None and carry_mask.any():
                                    _vis_rst[carry_mask] = (
                                        _vis_rst[carry_mask].astype(float) * 0.5
                                        + np.array([0, 0, 220]) * 0.5
                                    ).astype(np.uint8)
                                # 标注重置原因（取最后触发的那条）
                                _rst_reason = "RESET"
                                if _mask_vs_wok < 2.0:
                                    _rst_reason = f"RESET: mask过小({_mask_vs_wok:.1f}%<2%)"
                                elif last_reinforce_wok_pct > 5.0 and _drop_pct > 70.0:
                                    _rst_reason = f"RESET: 骤降({last_reinforce_wok_pct:.0f}%→{_mask_vs_wok:.0f}%)"
                                elif _mask_vs_wok > 35.0:
                                    _rst_reason = f"RESET: mask过大({_mask_vs_wok:.0f}%>wok35%)"
                                elif _overlap_pct < 60.0:
                                    _rst_reason = f"RESET: 偏离锅内(overlap={_overlap_pct:.0f}%)"
                                cv2.putText(_vis_rst, _rst_reason,
                                            (20, 40), cv2.FONT_HERSHEY_SIMPLEX,
                                            1.0, (0, 0, 255), 2)
                                cv2.putText(_vis_rst,
                                            f"t={chunk_start_s:.0f}s  f={chunk_start_abs}  [blue=bad_mask]",
                                            (20, 80), cv2.FONT_HERSHEY_SIMPLEX,
                                            0.75, (80, 80, 255), 2)
                                _rst_name = f"reset_t{chunk_start_s:.0f}s_f{chunk_start_abs}.jpg"
                                cv2.imwrite(os.path.join(out_dir, _rst_name), _vis_rst)
                                print(f"[重置] 预览图已保存: {_rst_name}  原因: {_rst_reason}")
                        except Exception as _rpe:
                            print(f"[重置] 预览图保存失败: {_rpe}")

                        carry_mask = None
                        last_relabel_s = chunk_start_s
                        # ── 尝试用 IR 当前帧低温区定位食材，生成新前景点 ────────
                        # 原理：食材温度 < 锅壁温度，取锅内低温像素 → 反投影到 RGB
                        _ir_fg_pts = []
                        if (temp_data is not None and homography is not None
                                and _wok_mask_al is not None and _H_inv_al is not None
                                and _rng_al is not None):
                            try:
                                _ir_idx_rst = min(int(chunk_start_abs * ir_fps_ratio),
                                                  temp_data.shape[0] - 1)
                                _ir_frm_rst = temp_data[_ir_idx_rst]
                                _wok_t_rst  = _ir_frm_rst[_wok_mask_al]
                                if len(_wok_t_rst) >= 10:
                                    # 取锅内温度低于 35 百分位的像素作为"食材候选"
                                    _t_thresh = float(np.percentile(_wok_t_rst, 35))
                                    _food_ir  = (_ir_frm_rst < _t_thresh) & _wok_mask_al
                                    _ys_ir, _xs_ir = np.where(_food_ir)
                                    if len(_xs_ir) >= 6:
                                        # 随机采 8 个 IR 像素
                                        _sel_ir = _rng_al.choice(len(_xs_ir),
                                                                  size=min(8, len(_xs_ir)),
                                                                  replace=False)
                                        # IR 坐标 → 齐次坐标 → 反投影到 RGB
                                        _pts_ir_h = np.stack([
                                            _xs_ir[_sel_ir].astype(float),
                                            _ys_ir[_sel_ir].astype(float),
                                            np.ones(len(_sel_ir))
                                        ])  # (3, N)
                                        _pts_rgb_h = _H_inv_al @ _pts_ir_h
                                        _pts_rgb_h = _pts_rgb_h[:2] / _pts_rgb_h[2]
                                        for _xi, _yi in _pts_rgb_h.T:
                                            _xi, _yi = float(_xi), float(_yi)
                                            if 0 <= _xi < VW and 0 <= _yi < VH:
                                                _ir_fg_pts.append([_xi, _yi])
                                        if _ir_fg_pts:
                                            print(f"[补强] t={chunk_start_s:.1f}s  "
                                                  f"重置+IR定位新前景点({len(_ir_fg_pts)}个)"
                                                  f"  阈值={_t_thresh:.1f}°C")
                            except Exception as _re:
                                print(f"[补强] IR定位失败({_re})，重置后用初始标注点")
                        if _ir_fg_pts:
                            # 把 IR 定位点作为下一批的补强注入（carry_mask=None 时走初始点路径，
                            # 但可以通过 inject_keyframes 在第 0 帧追加点）
                            _reinforce_inject = {
                                "local_frame": 0,
                                "fg_points":   _ir_fg_pts,
                                "bg_points":   [],
                                "label":       "IR-reset",
                            }
                    else:
                        # ── mask 正常，从内部腐蚀区域采点补强 ────────────────────
                        try:
                            _cm_u8 = carry_mask.astype(np.uint8) * 255
                            _ek = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (40, 40))
                            _inner = cv2.erode(_cm_u8, _ek)
                            _ys_r, _xs_r = np.where(_inner > 0)
                            _n_reinforce = 6
                            if len(_xs_r) >= _n_reinforce:
                                _idx_r = _rng_al.choice(len(_xs_r), size=_n_reinforce, replace=False)
                                _fg_reinforce = [[float(_xs_r[i]), float(_ys_r[i])] for i in _idx_r]
                            elif len(_xs_r) > 0:
                                _cx_r = float(np.mean(_xs_r)); _cy_r = float(np.mean(_ys_r))
                                _ys2, _xs2 = np.where(_cm_u8 > 0)
                                _dists = np.sqrt((_xs2-_cx_r)**2 + (_ys2-_cy_r)**2)
                                _sel = np.argsort(_dists)[:_n_reinforce]
                                _fg_reinforce = [[float(_xs2[i]), float(_ys2[i])] for i in _sel]
                            else:
                                _fg_reinforce = []

                            if _fg_reinforce:
                                _reinforce_inject = {
                                    "local_frame": 0,
                                    "fg_points":   _fg_reinforce,
                                    "bg_points":   [],
                                }
                                last_relabel_s = chunk_start_s
                                last_reinforce_wok_pct = _mask_vs_wok   # 记录本次补强时面积（骤降检测用）
                                print(f"[补强] t={chunk_start_s:.1f}s  mask正常({_mask_vs_wok:.0f}%)"
                                      f"  从SAM2末帧内部采点FG={len(_fg_reinforce)}")
                        except Exception as _e:
                            print(f"[补强] 采点失败({_e})，跳过本次补强")
                        # 保存预览图（不管采点是否成功，只要 carry_mask 有效就显示）
                        try:
                            _cap_al = cv2.VideoCapture(video_path)
                            _cap_al.set(cv2.CAP_PROP_POS_FRAMES, chunk_start_abs)
                            _ret_al, _rgb_al = _cap_al.read()
                            _cap_al.release()
                            if _ret_al:
                                _vis_al = _rgb_al.copy()
                                _vis_al[carry_mask] = (
                                    _vis_al[carry_mask].astype(float) * 0.6
                                    + np.array([0, 200, 80]) * 0.4
                                ).astype(np.uint8)
                                if _reinforce_inject:
                                    for _p in _reinforce_inject["fg_points"]:
                                        cv2.circle(_vis_al, (int(_p[0]), int(_p[1])), 8, (0, 255, 255), -1)
                                        cv2.circle(_vis_al, (int(_p[0]), int(_p[1])), 9, (0, 0, 0), 1)
                                n_pts = len(_reinforce_inject["fg_points"]) if _reinforce_inject else 0
                                cv2.putText(_vis_al,
                                            f"SAM2-reinforce t={chunk_start_s:.0f}s  pts(cyan):{n_pts}",
                                            (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 255), 2)
                                cv2.putText(_vis_al, "[carry_mask green | SAM2 mask preserved]",
                                            (20, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 200, 80), 2)
                                _preview_name = f"relabel_t{chunk_start_s:.0f}s_f{chunk_start_abs}.jpg"
                                cv2.imwrite(os.path.join(out_dir, _preview_name), _vis_al)
                                print(f"[补强] 预览图已保存: {_preview_name}")
                        except Exception as _pe:
                            print(f"[补强] 预览图保存失败: {_pe}")

            tmp_dir, frame_names, actual = extract_chunk_to_dir(
                video_path, chunk_start_abs, chunk_end_abs,
                infer_size=SAM2_INFER_SIZE
            )
            print(f"[抽帧] 临时目录: {tmp_dir}  实际帧数: {actual}")

            if actual == 0:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                break

            try:
                inject_this_chunk = []
                # 注入 food_labels.json 里的额外关键帧
                for abs_f, kf in inject_map.items():
                    if chunk_start_abs <= abs_f < chunk_end_abs:
                        local_f = abs_f - chunk_start_abs
                        inject_this_chunk.append({
                            "local_frame": local_f,
                            "fg_points":   kf["fg_points"],
                            "bg_points":   kf.get("bg_points", []),
                            "label":       kf.get("label", ""),
                        })
                # 注入补强点（local_frame=0，和 carry_mask 同帧，追加前景点）
                if _reinforce_inject is not None:
                    inject_this_chunk.append(_reinforce_inject)

                # carry_mask 需要缩放到推理分辨率传给下一批
                carry_mask_infer = None
                if carry_mask is not None and do_resize:
                    carry_mask_infer = upscale_mask(carry_mask, infer_wh) \
                        if carry_mask.shape != (infer_wh[1], infer_wh[0]) else carry_mask
                else:
                    carry_mask_infer = carry_mask

                chunk_masks, carry_mask_raw = track_chunk(
                    predictor, tmp_dir, frame_names,
                    fg_infer, bg_infer,
                    carry_mask=carry_mask_infer,
                    inject_keyframes=inject_this_chunk,
                )
                # carry_mask 放大回原始分辨率，供下批 add_new_mask 使用
                if do_resize and carry_mask_raw is not None:
                    carry_mask = upscale_mask(carry_mask_raw, orig_wh)
                else:
                    carry_mask = carry_mask_raw
                # ── 对 carry_mask 也做 wok 约束，防止面积检查出现 >100% ──────
                if carry_mask is not None and _wok_rgb_constraint is not None:
                    carry_mask = carry_mask & _wok_rgb_constraint
                print(f"[SAM2] 批次 {chunk_i+1} 追踪完成，{len(chunk_masks)} 帧")

                # ── mask 面积异常检测：若本批平均占比 > 60% 且比上批大 3 倍 → 强制下批重标点 ──
                if len(chunk_masks) > 0:
                    _total_px = infer_wh[0] * infer_wh[1] if do_resize else VW * VH
                    _ratios   = [m.sum() / _total_px * 100
                                 for m in chunk_masks.values() if m is not None]
                    _mean_ratio = float(np.mean(_ratios)) if _ratios else 0.0
                    if (_auto_label_func is not None
                            and _mean_ratio > 60.0
                            and _mean_ratio > prev_mask_ratio * 3.0
                            and prev_mask_ratio > 0.5):
                        carry_mask = None
                        last_relabel_s = -999   # 强制下批立即重标点
                        print(f"[异常检测] 批次{chunk_i+1} 平均mask={_mean_ratio:.1f}%"
                              f"（上批{prev_mask_ratio:.1f}%），追踪失控！强制下批重标点")
                    prev_mask_ratio = _mean_ratio
            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                print(f"[清理] 临时帧已删除")

            cap = cv2.VideoCapture(video_path)
            cap.set(cv2.CAP_PROP_POS_FRAMES, chunk_start_abs)

            for local_in_chunk in range(actual):
                ret, frame = cap.read()
                if not ret:
                    break

                abs_idx   = chunk_start_abs + local_in_chunk
                local_idx = global_local + local_in_chunk
                # SAM2 返回的是推理分辨率的 mask，放大回原始分辨率
                raw_mask  = chunk_masks.get(local_in_chunk,
                                            np.zeros((infer_wh[1], infer_wh[0]), dtype=bool))
                mask      = upscale_mask(raw_mask, orig_wh) if do_resize else raw_mask
                mask_src  = "SAM2"

                # ── wok 约束后处理：mask AND 预计算的锅内区域 ────────────────
                if _wok_rgb_constraint is not None:
                    mask = mask & _wok_rgb_constraint

                vis       = render_overlay(frame, mask, MASK_COLOR, MASK_ALPHA)
                temp_mean = temp_min = temp_max = float("nan")
                if temp_data is not None and homography is not None:
                    ir_h, ir_w = temp_data.shape[1], temp_data.shape[2]
                    ir_mask    = map_mask_to_ir(mask, homography, (ir_h, ir_w))
                    ir_idx     = int(abs_idx * ir_fps_ratio)
                    if ir_idx < temp_data.shape[0]:
                        t_frame    = temp_data[ir_idx]
                        food_temps = t_frame[ir_mask]
                        if len(food_temps) > 0:
                            temp_mean = float(np.mean(food_temps))
                            temp_min  = float(np.min(food_temps))
                            temp_max  = float(np.max(food_temps))

                time_s     = abs_idx / fps
                mask_ratio = mask.sum() / mask.size * 100
                if not np.isnan(temp_mean):
                    temp_history.append((time_s, temp_mean))

                # ── ROI 温度统计 ──────────────────────────────────────────────
                roi_temp_mean = float("nan")
                if roi_cfg is not None and temp_data is not None and homography is not None:
                    try:
                        ir_h, ir_w = temp_data.shape[1], temp_data.shape[2]
                        ir_idx_roi = int(abs_idx * ir_fps_ratio)
                        if ir_idx_roi < temp_data.shape[0]:
                            # RGB 圆心 → IR 坐标
                            rgb_pt = np.array([[[float(roi_cfg["rgb_cx"]),
                                                 float(roi_cfg["rgb_cy"])]]], dtype=np.float32)
                            ir_pt  = cv2.perspectiveTransform(rgb_pt, homography)[0][0]
                            ir_cx_roi = int(round(ir_pt[0]))
                            ir_cy_roi = int(round(ir_pt[1]))
                            # RGB 半径 → IR 半径（按 IR/RGB 宽度比缩放）
                            ir_r = max(1, int(roi_cfg["rgb_radius"] * ir_w / roi_cfg["rgb_w"]))
                            # 生成圆形 mask
                            roi_mask_ir = np.zeros((ir_h, ir_w), dtype=np.uint8)
                            cv2.circle(roi_mask_ir, (ir_cx_roi, ir_cy_roi), ir_r, 255, -1)
                            roi_temps = temp_data[ir_idx_roi][roi_mask_ir > 0]
                            if len(roi_temps) > 0:
                                roi_temp_mean = float(np.mean(roi_temps))
                    except Exception:
                        pass
                if not np.isnan(roi_temp_mean):
                    roi_history.append((time_s, roi_temp_mean))

                # 在 RGB 帧上叠加 ROI 圆形
                if roi_cfg is not None:
                    cv2.circle(vis, (roi_cfg["rgb_cx"], roi_cfg["rgb_cy"]),
                               roi_cfg["rgb_radius"], (255, 200, 0), 2)
                    if not np.isnan(roi_temp_mean):
                        cv2.putText(vis, f"ROI:{roi_temp_mean:.1f}C",
                                    (roi_cfg["rgb_cx"] - 40,
                                     roi_cfg["rgb_cy"] - roi_cfg["rgb_radius"] - 8),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 200, 0), 1)

                info_bar = np.zeros((INFO_H, VW, 3), dtype=np.uint8)
                cv2.putText(info_bar,
                            f"Frame {abs_idx}  t={time_s:.1f}s  Mask={mask_ratio:.1f}%  [SAM2]",
                            (10, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)

                # ── IR mask 自动分割温度（K-means 双峰分类）──────────────────
                # 锅内温度分布是双峰：高温峰=锅壁/锅底，低温峰=食物
                # K-means 自适应分类，比固定百分位更准确
                ir_mask_temp = float("nan")
                if wok_mask_ir is not None and temp_data is not None:
                    ir_idx_wok = int(abs_idx * ir_fps_ratio)
                    if ir_idx_wok < temp_data.shape[0]:
                        t_frame_wok = temp_data[ir_idx_wok]
                        wok_temps   = t_frame_wok[wok_mask_ir]
                        if len(wok_temps) >= 10:
                            ir_mask_temp = _kmeans_food_temp(wok_temps)
                if not np.isnan(ir_mask_temp):
                    ir_mask_history.append((time_s, ir_mask_temp))

                # ── info_bar 第二行：三个温度值 ───────────────────────────────
                parts = []
                if not np.isnan(temp_mean):
                    parts.append(f"SAM2:{temp_mean:.1f}C(min{temp_min:.0f}/max{temp_max:.0f})")
                else:
                    parts.append("SAM2:N/A")
                if not np.isnan(roi_temp_mean):
                    parts.append(f"ROI:{roi_temp_mean:.1f}C")
                if not np.isnan(ir_mask_temp):
                    parts.append(f"IR:{ir_mask_temp:.1f}C")
                cv2.putText(info_bar, "  ".join(parts),
                            (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (100, 255, 200), 1)

                chart_bar = draw_temp_chart(temp_history, time_s, VW, CHART_H, CURVE_WIN_S,
                                            roi_history=roi_history,
                                            ir_mask_history=ir_mask_history)
                writer.write(np.vstack([vis, info_bar, chart_bar]))
                # 三策略数据分别记录
                sam2_rows.append([abs_idx, local_idx, time_s,
                                  int(mask.sum()), round(mask_ratio, 2),
                                  temp_mean, temp_min, temp_max])
                roi_rows.append([abs_idx, local_idx, time_s, roi_temp_mean])
                ir_rows.append([abs_idx, local_idx, time_s, ir_mask_temp])
                if local_in_chunk % 50 == 0:
                    print(f"  写帧: {local_in_chunk+1}/{actual}  "
                          f"mask={mask_ratio:.1f}%  temp={temp_mean:.1f}°C", end="\r")

            cap.release()
            print()
            global_local += actual

    else:
        # ── 混合模式：逐帧处理，锚点帧用 SAM2，中间帧用光流 ──────────────────
        cap = cv2.VideoCapture(video_path)
        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        anchor_set = set(anchor_frames)   # 快速查找
        n_total    = len(anchor_frames)
        anchor_cnt = 0

        for abs_idx in range(start_frame, total_frames):
            ret, frame = cap.read()
            if not ret:
                break

            local_idx  = abs_idx - start_frame
            is_anchor  = (abs_idx in anchor_set)

            if is_anchor:
                # SAM2 单帧推理：抽单帧到临时目录
                tmp_dir, frame_names, actual = extract_chunk_to_dir(
                    video_path, abs_idx, abs_idx + 1
                )
                try:
                    # 检查此帧是否有注入关键帧（对于第一帧之外的额外标注）
                    inject_this = []
                    if abs_idx in inject_map and abs_idx != start_frame:
                        kf = inject_map[abs_idx]
                        inject_this = [{
                            "local_frame": 0,
                            "fg_points":   kf["fg_points"],
                            "bg_points":   kf.get("bg_points", []),
                            "label":       kf.get("label", ""),
                        }]

                    chunk_masks, carry_mask = track_chunk(
                        predictor, tmp_dir, frame_names,
                        fg_points, bg_points,
                        carry_mask=carry_mask,
                        inject_keyframes=inject_this,
                    )
                    mask     = chunk_masks.get(0, np.zeros((VH, VW), dtype=bool))
                    mask_src = "SAM2"
                    anchor_cnt += 1
                    if anchor_cnt % 25 == 1 or anchor_cnt == n_total:
                        print(f"  [SAM2锚点 {anchor_cnt}/{n_total}] 帧 {abs_idx}"
                              f"  t={abs_idx/fps:.1f}s", end="\r")
                finally:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
            else:
                # 光流传播
                if flow_prev_gray is None or flow_prev_mask is None:
                    # 兜底：没有前帧信息时用空 mask
                    mask = np.zeros((VH, VW), dtype=bool)
                else:
                    cur_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    mask     = flow_propagate_mask(flow_prev_gray, cur_gray, flow_prev_mask)
                mask_src = "Flow"

            # 更新光流状态
            flow_prev_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            flow_prev_mask = mask

            # 温度统计
            temp_mean = temp_min = temp_max = float("nan")
            if temp_data is not None and homography is not None:
                ir_h, ir_w = temp_data.shape[1], temp_data.shape[2]
                ir_mask    = map_mask_to_ir(mask, homography, (ir_h, ir_w))
                ir_idx     = int(abs_idx * ir_fps_ratio)
                if ir_idx < temp_data.shape[0]:
                    t_frame    = temp_data[ir_idx]
                    food_temps = t_frame[ir_mask]
                    if len(food_temps) > 0:
                        temp_mean = float(np.mean(food_temps))
                        temp_min  = float(np.min(food_temps))
                        temp_max  = float(np.max(food_temps))

            time_s     = abs_idx / fps
            mask_ratio = mask.sum() / mask.size * 100
            if not np.isnan(temp_mean):
                temp_history.append((time_s, temp_mean))

            info_bar  = np.zeros((INFO_H, VW, 3), dtype=np.uint8)
            hud_color = (255, 255, 255) if mask_src == "SAM2" else (80, 220, 255)
            color     = MASK_COLOR if mask_src == "SAM2" else (0, 200, 80)
            vis       = render_overlay(frame, mask, color, MASK_ALPHA)

            cv2.putText(info_bar,
                        f"Frame {abs_idx}  t={time_s:.1f}s  "
                        f"Mask={mask_ratio:.1f}%  [{mask_src}]",
                        (10, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.55, hud_color, 1)
            cv2.putText(info_bar,
                        (f"Temp: mean={temp_mean:.1f}C  min={temp_min:.1f}C  max={temp_max:.1f}C"
                         if not np.isnan(temp_mean) else "Temp: N/A"),
                        (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (100, 255, 200), 1)

            chart_bar = draw_temp_chart(temp_history, time_s, VW, CHART_H, CURVE_WIN_S)
            writer.write(np.vstack([vis, info_bar, chart_bar]))
            sam2_rows.append([abs_idx, local_idx, time_s,
                              int(mask.sum()), round(mask_ratio, 2),
                              temp_mean, temp_min, temp_max])

        cap.release()
        print()
        global_local = total_frames - start_frame

    writer.release()

    print(f"\n\n✅ 追踪完成！")
    print(f"   结果目录:   {out_dir}")
    print(f"   可视化视频: {out_video_viz}")
    print(f"   共处理 {global_local} 帧")

    # ── 保存三策略独立 Excel ──────────────────────────────────────────────────
    _save_three_xlsx(sam2_rows, roi_rows, ir_rows, out_dir)

    # ── 绘制三策略温度曲线 PNG ────────────────────────────────────────────────
    _plot_three_curves(sam2_rows, roi_rows, ir_rows, out_curve)

    # ── 拼合 RGB + IR 视频 ────────────────────────────────────────────────────
    if temp_data is not None and wok_cfg is not None:
        out_combined = os.path.join(out_dir, "track_result_combined.mp4")
        ir_fps_val   = fps * ir_fps_ratio   # 估算 IR 帧率
        print(f"\n[拼合] 开始生成 RGB+IR 并排视频...")
        stitch_rgb_ir(
            rgb_viz_path=out_video_viz,
            temp_data=temp_data,
            ir_fps=ir_fps_val,
            wok_cfg=wok_cfg,
            out_path=out_combined,
            rgb_start_frame=start_frame,
            rgb_fps=fps,
            pct=IR_FOOD_PCT,
        )
        print(f"   并排视频: {out_combined}")
        # 删除中间产物（纯 RGB viz 视频），只保留最终并排视频
        try:
            os.remove(out_video_viz)
            print(f"   已删除中间文件: track_result_viz.mp4")
        except Exception:
            pass
    else:
        print("[拼合] 缺少温度数据或锅区域配置，跳过 IR 拼合")

    # 清理空的 tmp 根目录（如果为空）
    try:
        if os.path.exists(TMP_BASE) and not os.listdir(TMP_BASE):
            os.rmdir(TMP_BASE)
    except Exception:
        pass


def _save_excel(csv_path, xlsx_path):
    """从 CSV 生成 Excel，包含逐帧数据和汇总统计两个 sheet"""
    if not _HAS_OPENPYXL:
        print("[Excel] 未安装 openpyxl，跳过。运行: pip install openpyxl")
        return
    import csv as csv_mod

    wb = openpyxl.Workbook()

    # ── Sheet1：逐帧数据 ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "逐帧数据"

    headers = ["帧号(绝对)", "帧号(相对)", "时间(s)", "Mask像素数",
               "Mask占比(%)", "温度均值(°C)", "温度最小(°C)", "温度最大(°C)"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    rows_data = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv_mod.DictReader(f)
        for row in reader:
            try:
                rows_data.append([
                    int(row["frame_abs"]),
                    int(row["frame_rel"]),
                    round(float(row["time_s"]), 3),
                    int(row["mask_pixels"]),
                    round(float(row["mask_ratio"]), 2),
                    round(float(row["temp_mean"]), 2) if row["temp_mean"] != "nan" else None,
                    round(float(row["temp_min"]),  2) if row["temp_min"]  != "nan" else None,
                    round(float(row["temp_max"]),  2) if row["temp_max"]  != "nan" else None,
                ])
            except (ValueError, KeyError):
                continue

    for r_idx, row_vals in enumerate(rows_data, 2):
        for c_idx, val in enumerate(row_vals, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)

    # 列宽自适应
    col_widths = [12, 12, 10, 14, 12, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet2：汇总统计 ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("汇总统计")
    valid_temps = [r[5] for r in rows_data if r[5] is not None]
    valid_times = [r[2] for r in rows_data if r[5] is not None]

    stats = []
    if valid_temps:
        stats = [
            ("总追踪帧数",    len(rows_data)),
            ("有效温度帧数",  len(valid_temps)),
            ("追踪时长(s)",   round(rows_data[-1][2] - rows_data[0][2], 1) if rows_data else 0),
            ("温度均值(°C)",  round(float(np.mean(valid_temps)), 2)),
            ("温度最大(°C)",  round(float(np.max(valid_temps)),  2)),
            ("温度最小(°C)",  round(float(np.min(valid_temps)),  2)),
            ("温度标准差",    round(float(np.std(valid_temps)),   2)),
            ("峰值温度时刻(s)", round(valid_times[int(np.argmax(valid_temps))], 1)),
        ]
    else:
        stats = [("总追踪帧数", len(rows_data)), ("有效温度帧数", 0)]

    ws2.cell(row=1, column=1, value="统计项").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="数值").font   = Font(bold=True)
    for i, (k, v) in enumerate(stats, 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    wb.save(xlsx_path)
    print(f"[Excel] 已保存: {xlsx_path}  ({len(rows_data)} 行数据)")


def _plot_temp_curve(csv_path, out_path=None):
    """从 CSV 绘制菜温随时间变化曲线"""
    import csv
    times, means, mins, maxs = [], [], [], []
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                tm = float(row["temp_mean"])
                if not np.isnan(tm):
                    times.append(float(row["time_s"]))
                    means.append(tm)
                    mins.append(float(row["temp_min"]))
                    maxs.append(float(row["temp_max"]))
            except (ValueError, KeyError):
                continue

    if not times:
        print("[温度曲线] 无有效温度数据，跳过绘图")
        return

    fig, ax = plt.subplots(figsize=(12, 4))
    ax.fill_between(times, mins, maxs, alpha=0.2, color="orange", label="min~max range")
    ax.plot(times, means, color="red", linewidth=1.5, label="mean temp")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Temp (C)")
    ax.set_title("Food Temperature over Time (SAM2 Mask Region)")
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    out = out_path or os.path.join(_HERE, "..", "output", "food_temp_curve.png")
    plt.savefig(out, dpi=120)
    plt.close()
    print(f"[温度曲线] 已保存: {out}")


# ── 三策略数据保存 ────────────────────────────────────────────────────────────

def _save_three_xlsx(sam2_rows, roi_rows, ir_rows, out_dir):
    """
    输出三个独立 xlsx 文件，每个对应一种温度策略：
      temp_sam2.xlsx  — SAM2 mask 追踪温度
      temp_roi.xlsx   — ROI 固定圆圈温度
      temp_ir.xlsx    — IR 自动分割（锅内低温区）温度
    """
    if not _HAS_OPENPYXL:
        print("[Excel] 未安装 openpyxl，跳过。pip install openpyxl")
        return

    def _make_wb(headers, rows, fill_color):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "逐帧数据"
        hdr_fill = PatternFill("solid", fgColor=fill_color)
        hdr_font = Font(bold=True, color="FFFFFF")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        valid_rows = []
        for r_idx, row in enumerate(rows, 2):
            clean = [v if not isinstance(v, float) or not np.isnan(v)
                     else None for v in row]
            for c_idx, val in enumerate(clean, 1):
                ws.cell(row=r_idx, column=c_idx,
                        value=round(val, 3) if isinstance(val, float) else val)
            valid_rows.append(clean)
        # 汇总 sheet
        ws2 = wb.create_sheet("汇总统计")
        temps = [r[-1] for r in valid_rows
                 if r[-1] is not None and not np.isnan(r[-1])]
        times = [r[2] for r in valid_rows
                 if r[-1] is not None and not np.isnan(r[-1])]
        if temps:
            stats = [
                ("总帧数",        len(rows)),
                ("有效温度帧数",  len(temps)),
                ("追踪时长(s)",   round(rows[-1][2]-rows[0][2], 1) if rows else 0),
                ("温度均值(°C)",  round(float(np.mean(temps)), 2)),
                ("温度最大(°C)",  round(float(np.max(temps)),  2)),
                ("温度最小(°C)",  round(float(np.min(temps)),  2)),
                ("温度标准差",    round(float(np.std(temps)),   2)),
                ("峰值时刻(s)",   round(times[int(np.argmax(temps))], 1)),
            ]
        else:
            stats = [("总帧数", len(rows)), ("有效温度帧数", 0)]
        ws2.cell(row=1, column=1, value="统计项").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="数值").font   = Font(bold=True)
        for i, (k, v) in enumerate(stats, 2):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 14
        return wb, len(valid_rows)

    # SAM2
    h1 = ["帧号(绝对)","帧号(相对)","时间(s)","Mask像素","Mask%",
          "均值(°C)","最小(°C)","最大(°C)"]
    wb1, n1 = _make_wb(h1, sam2_rows, "1F4E79")
    p1 = os.path.join(out_dir, "temp_sam2.xlsx")
    wb1.save(p1)
    print(f"[Excel] SAM2  → {p1}  ({n1} 行)")

    # ROI
    h2 = ["帧号(绝对)","帧号(相对)","时间(s)","ROI均值(°C)"]
    wb2, n2 = _make_wb(h2, roi_rows, "833C00")
    p2 = os.path.join(out_dir, "temp_roi.xlsx")
    wb2.save(p2)
    print(f"[Excel] ROI   → {p2}  ({n2} 行)")

    # IR
    h3 = ["帧号(绝对)","帧号(相对)","时间(s)","IR自动均值(°C)"]
    wb3, n3 = _make_wb(h3, ir_rows, "375623")
    p3 = os.path.join(out_dir, "temp_ir.xlsx")
    wb3.save(p3)
    print(f"[Excel] IR    → {p3}  ({n3} 行)")


def _plot_three_curves(sam2_rows, roi_rows, ir_rows, out_path):
    """绘制三条温度曲线对比图（matplotlib）"""
    def _extract(rows, val_col):
        xs, ys = [], []
        for r in rows:
            v = r[val_col]
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                xs.append(r[2])   # time_s
                ys.append(float(v))
        return xs, ys

    fig, ax = plt.subplots(figsize=(14, 4))

    t_sam2, v_sam2 = _extract(sam2_rows, 5)   # temp_mean
    t_roi,  v_roi  = _extract(roi_rows,  3)   # roi_temp_mean
    t_ir,   v_ir   = _extract(ir_rows,   3)   # ir_mask_temp

    if t_sam2:
        # min/max 带
        t_min = [r[6] for r in sam2_rows if r[6] is not None and not np.isnan(r[6])]
        t_max = [r[7] for r in sam2_rows if r[7] is not None and not np.isnan(r[7])]
        if t_min and t_max and len(t_min) == len(t_sam2):
            ax.fill_between(t_sam2, t_min, t_max, alpha=0.15, color="#FF7F0E")
        ax.plot(t_sam2, v_sam2, color="#FF7F0E", lw=1.5, label="SAM2 Mask")
    if t_roi:
        ax.plot(t_roi,  v_roi,  color="#1F77B4", lw=1.5, label="ROI Fixed")
    if t_ir:
        ax.plot(t_ir,   v_ir,   color="#2CA02C", lw=1.5, label="IR Auto")

    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Temperature (°C)")
    ax.set_title("Food Temperature — Three Strategies Comparison")
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"[温度曲线] 已保存: {out_path}")


# ── RGB + IR 视频拼合 ─────────────────────────────────────────────────────────

def stitch_rgb_ir(rgb_viz_path, temp_data, ir_fps, wok_cfg,
                  out_path, rgb_start_frame, rgb_fps, pct=40):
    """
    新布局：
      ┌─────────────────────────────────────────────────┐
      │  RGB纯视频（带mask）  │  IR热图（等高，上下对齐）  │
      ├─────────────────────────────────────────────────┤
      │       INFO BAR — 三种策略温度数值（全宽，放大）    │
      ├─────────────────────────────────────────────────┤
      │       温度曲线图（三条曲线，全宽，放大）           │
      └─────────────────────────────────────────────────┘

    从已生成的 track_result_viz.mp4 里裁取：
      - 纯视频区：rows 0 .. VH-1
      - info_bar ：rows VH .. VH+INFO_H-1
      - chart    ：rows VH+INFO_H .. end
    然后按新布局拼合。
    """
    # 动态导入，避免循环依赖
    import sys as _sys
    _tools = os.path.join(_HERE, "..", "tools")
    if _tools not in _sys.path:
        _sys.path.insert(0, _tools)
    from ir_mask_viz import render_ir_frame

    cap = cv2.VideoCapture(rgb_viz_path)
    if not cap.isOpened():
        print(f"[拼合] 无法打开 RGB 视频: {rgb_viz_path}")
        return

    rgb_total  = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    out_fps    = cap.get(cv2.CAP_PROP_FPS) or rgb_fps
    viz_w      = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))   # viz 视频宽（= VW = 1600）
    viz_h      = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))  # viz 视频高（= VH+INFO_H+CHART_H）

    # viz 里各区域高度（与 main() 里的常量保持一致）
    pure_h  = viz_h - INFO_H - CHART_H   # 纯视频高度（= VH = 1200）
    pure_w  = viz_w                       # 纯视频宽度（= VW = 1600）

    n_ir = temp_data.shape[0]

    # IR 渲染尺寸：等比缩放到与纯视频同高（pure_h），上下对齐
    ir_aspect = temp_data.shape[2] / temp_data.shape[1]   # ir_w / ir_h
    ir_out_h  = pure_h
    ir_out_w  = int(round(ir_out_h * ir_aspect))
    if ir_out_w % 2 != 0:
        ir_out_w += 1

    # 合并视频总宽度 = RGB宽 + IR宽
    total_w = pure_w + ir_out_w

    # 底部 info + chart 放大高度（比原来更高，视觉更清晰）
    new_info_h  = int(INFO_H  * 1.8)   # ≈ 90px
    new_chart_h = int(CHART_H * 1.8)   # ≈ 216px
    if new_info_h  % 2 != 0: new_info_h  += 1
    if new_chart_h % 2 != 0: new_chart_h += 1

    total_h = pure_h + new_info_h + new_chart_h

    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(out_path, fourcc, out_fps, (total_w, total_h))

    print(f"[拼合] 纯视频: RGB={pure_w}×{pure_h}  IR={ir_out_w}×{ir_out_h}")
    print(f"[拼合] 底部面板: info={new_info_h}px  chart={new_chart_h}px  全宽={total_w}px")
    print(f"[拼合] 输出帧尺寸: {total_w}×{total_h}  共 {rgb_total} 帧")

    for idx in range(rgb_total):
        ret, viz_frame = cap.read()
        if not ret:
            break

        # ── 裁取各区域 ───────────────────────────────────────────────────────
        pure_rgb  = viz_frame[0:pure_h, :]                           # 纯视频帧
        info_src  = viz_frame[pure_h:pure_h + INFO_H, :]             # info bar（原始小）
        chart_src = viz_frame[pure_h + INFO_H:pure_h + INFO_H + CHART_H, :]  # chart（原始小）

        # ── 渲染 IR 帧（等高对齐）─────────────────────────────────────────────
        time_s  = (rgb_start_frame + idx) / rgb_fps
        ir_idx  = min(int(round(time_s * ir_fps)), n_ir - 1)
        ir_img  = render_ir_frame(
            temp_data[ir_idx], wok_cfg, pct=pct,
            out_w=ir_out_w, out_h=ir_out_h
        )

        # ── 上方：RGB | IR 并排（等高，上下对齐）─────────────────────────────
        top_panel = np.hstack([pure_rgb, ir_img])   # (pure_h, total_w, 3)

        # ── 底部 info bar：放大到全宽 ──────────────────────────────────────────
        # 先在 info_src 两侧填充黑色，扩展到 total_w，再放大高度
        # 简单做法：直接 resize 到 (total_w, new_info_h)
        info_big = cv2.resize(info_src, (total_w, new_info_h),
                              interpolation=cv2.INTER_LINEAR)

        # ── 底部 chart：同样放大到全宽 ─────────────────────────────────────────
        chart_big = cv2.resize(chart_src, (total_w, new_chart_h),
                               interpolation=cv2.INTER_LINEAR)

        # ── 拼合 ──────────────────────────────────────────────────────────────
        combined = np.vstack([top_panel, info_big, chart_big])
        writer.write(combined)

        if idx % 200 == 0:
            print(f"  拼合帧 {idx}/{rgb_total}  IR帧={ir_idx}", end="\r")

    cap.release()
    writer.release()
    print(f"\n[拼合] 完成: {out_path}")


if __name__ == "__main__":
    main()
